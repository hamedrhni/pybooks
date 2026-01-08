# exporters/__init__.py
# Copyright (C) 2024 - 2028 the PythonAccounting authors and contributors
# <see AUTHORS file>
#
# This module is part of PythonAccounting and is released under
# the MIT License: https://www.opensource.org/licenses/mit-license.php

"""
Export functionality for financial reports.

This module provides exporters for various formats including CSV, JSON, Excel, and PDF.
"""

from __future__ import annotations

from abc import ABC, abstractmethod
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any, Dict, List, Optional, Union

if TYPE_CHECKING:
    from python_accounting.reports import FinancialStatement


class BaseExporter(ABC):
    """Base class for all exporters."""
    
    def __init__(self, report: "FinancialStatement"):
        """
        Initialize the exporter.
        
        Args:
            report: The financial statement to export.
        """
        self.report = report
        self.entity_name = report.session.entity.name
        self.report_title = report.title
        self.generated_at = datetime.now()
    
    @abstractmethod
    def export(self, filepath: Union[str, Path]) -> Path:
        """
        Export the report to a file.
        
        Args:
            filepath: The path to save the exported file.
            
        Returns:
            The path to the exported file.
        """
        pass
    
    @abstractmethod
    def to_dict(self) -> Dict[str, Any]:
        """
        Convert the report to a dictionary.
        
        Returns:
            Dictionary representation of the report.
        """
        pass
    
    def _get_sections_data(self) -> List[Dict[str, Any]]:
        """Extract sections data from the report."""
        sections_data = []
        
        for section in self.report.sections:
            section_info = {
                "name": section.value,
                "accounts": {},
                "balances": self.report.balances.get(section.name, {}),
                "total": self.report.totals.get(section.name, 0),
            }
            
            if section.name in self.report.accounts:
                section_info["accounts"] = self.report.accounts[section.name]
            
            sections_data.append(section_info)
        
        return sections_data


class CSVExporter(BaseExporter):
    """Export reports to CSV format."""
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert report to dictionary."""
        return {
            "entity": self.entity_name,
            "report": self.report_title,
            "generated_at": self.generated_at.isoformat(),
            "sections": self._get_sections_data(),
            "results": self.report.result_amounts,
        }
    
    def export(self, filepath: Union[str, Path]) -> Path:
        """
        Export the report to CSV.
        
        Args:
            filepath: The path to save the CSV file.
            
        Returns:
            The path to the exported file.
        """
        import csv
        
        filepath = Path(filepath)
        
        with open(filepath, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            
            # Header
            writer.writerow([self.entity_name])
            writer.writerow([self.report_title])
            writer.writerow([f"Generated: {self.generated_at.strftime('%Y-%m-%d %H:%M:%S')}"])
            writer.writerow([])
            
            # Sections
            for section in self.report.sections:
                writer.writerow([section.value])
                
                section_balances = self.report.balances.get(section.name, {})
                for account_type, balance in section_balances.items():
                    writer.writerow(["", account_type, balance])
                
                writer.writerow(
                    ["", f"Total {section.value}", self.report.totals.get(section.name, 0)]
                )
                writer.writerow([])
            
            # Results
            writer.writerow(["Results"])
            for result_name, amount in self.report.result_amounts.items():
                writer.writerow(["", result_name, amount])
        
        return filepath


class JSONExporter(BaseExporter):
    """Export reports to JSON format."""
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert report to dictionary."""
        sections = {}
        for section in self.report.sections:
            sections[section.name] = {
                "label": section.value,
                "accounts": self.report.accounts.get(section.name, {}),
                "balances": {
                    k: float(v) for k, v in self.report.balances.get(section.name, {}).items()
                },
                "total": float(self.report.totals.get(section.name, 0)),
            }
        
        return {
            "metadata": {
                "entity": self.entity_name,
                "report_type": self.report_title,
                "generated_at": self.generated_at.isoformat(),
                "start_date": getattr(self.report, "start_date", None),
                "end_date": getattr(self.report, "end_date", None),
            },
            "sections": sections,
            "results": {k: float(v) for k, v in self.report.result_amounts.items()},
            "totals": {
                "debit": float(self.report.balances.get("debit", 0)),
                "credit": float(self.report.balances.get("credit", 0)),
            },
        }
    
    def export(self, filepath: Union[str, Path]) -> Path:
        """
        Export the report to JSON.
        
        Args:
            filepath: The path to save the JSON file.
            
        Returns:
            The path to the exported file.
        """
        import json
        
        filepath = Path(filepath)
        
        data = self.to_dict()
        
        # Convert datetime objects to strings
        if data["metadata"]["start_date"]:
            data["metadata"]["start_date"] = data["metadata"]["start_date"].isoformat()
        if data["metadata"]["end_date"]:
            data["metadata"]["end_date"] = data["metadata"]["end_date"].isoformat()
        
        with open(filepath, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, default=str)
        
        return filepath


class ExcelExporter(BaseExporter):
    """Export reports to Excel format."""
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert report to dictionary."""
        return JSONExporter(self.report).to_dict()
    
    def export(self, filepath: Union[str, Path]) -> Path:
        """
        Export the report to Excel.
        
        Args:
            filepath: The path to save the Excel file.
            
        Returns:
            The path to the exported file.
            
        Raises:
            ImportError: If openpyxl is not installed.
        """
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install python-accounting[excel]"
            )
        
        filepath = Path(filepath)
        
        wb = Workbook()
        ws = wb.active
        ws.title = self.report_title[:31]  # Excel sheet name limit
        
        # Styles
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=12)
        bold_font = Font(bold=True)
        thin_border = Border(bottom=Side(style="thin"))
        
        row = 1
        
        # Header
        ws.cell(row=row, column=1, value=self.entity_name).font = title_font
        row += 1
        ws.cell(row=row, column=1, value=self.report_title).font = header_font
        row += 1
        ws.cell(
            row=row, column=1, 
            value=f"Generated: {self.generated_at.strftime('%Y-%m-%d %H:%M:%S')}"
        )
        row += 2
        
        # Sections
        for section in self.report.sections:
            ws.cell(row=row, column=1, value=section.value).font = bold_font
            row += 1
            
            section_balances = self.report.balances.get(section.name, {})
            for account_type, balance in section_balances.items():
                ws.cell(row=row, column=2, value=account_type)
                ws.cell(row=row, column=3, value=float(balance))
                row += 1
            
            ws.cell(row=row, column=2, value=f"Total {section.value}").font = bold_font
            cell = ws.cell(row=row, column=3, value=float(self.report.totals.get(section.name, 0)))
            cell.font = bold_font
            cell.border = thin_border
            row += 2
        
        # Results
        for result_name, amount in self.report.result_amounts.items():
            ws.cell(row=row, column=1, value=result_name).font = bold_font
            ws.cell(row=row, column=3, value=float(amount)).font = bold_font
            row += 1
        
        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 30
        ws.column_dimensions["C"].width = 15
        
        wb.save(filepath)
        return filepath


class PDFExporter(BaseExporter):
    """Export reports to PDF format."""
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert report to dictionary."""
        return JSONExporter(self.report).to_dict()
    
    def export(self, filepath: Union[str, Path]) -> Path:
        """
        Export the report to PDF.
        
        Args:
            filepath: The path to save the PDF file.
            
        Returns:
            The path to the exported file.
            
        Raises:
            ImportError: If reportlab is not installed.
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import letter, A4
            from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
            )
        except ImportError:
            raise ImportError(
                "reportlab is required for PDF export. "
                "Install it with: pip install python-accounting[pdf]"
            )
        
        filepath = Path(filepath)
        
        doc = SimpleDocTemplate(str(filepath), pagesize=A4)
        styles = getSampleStyleSheet()
        elements = []
        
        # Title
        title_style = ParagraphStyle(
            "CustomTitle",
            parent=styles["Heading1"],
            alignment=1,  # Center
        )
        elements.append(Paragraph(self.entity_name, title_style))
        elements.append(Paragraph(self.report_title, styles["Heading2"]))
        elements.append(
            Paragraph(
                f"Generated: {self.generated_at.strftime('%Y-%m-%d %H:%M:%S')}",
                styles["Normal"]
            )
        )
        elements.append(Spacer(1, 0.5 * inch))
        
        # Build table data
        table_data = []
        
        for section in self.report.sections:
            table_data.append([section.value, "", ""])
            
            section_balances = self.report.balances.get(section.name, {})
            for account_type, balance in section_balances.items():
                table_data.append(["", account_type, f"{float(balance):,.2f}"])
            
            table_data.append(
                ["", f"Total {section.value}", 
                 f"{float(self.report.totals.get(section.name, 0)):,.2f}"]
            )
            table_data.append(["", "", ""])
        
        # Results
        for result_name, amount in self.report.result_amounts.items():
            table_data.append([result_name, "", f"{float(amount):,.2f}"])
        
        # Create table
        table = Table(table_data, colWidths=[2.5 * inch, 2.5 * inch, 1.5 * inch])
        table.setStyle(TableStyle([
            ("FONTNAME", (0, 0), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 0), (-1, -1), 10),
            ("ALIGN", (2, 0), (2, -1), "RIGHT"),
            ("TOPPADDING", (0, 0), (-1, -1), 3),
            ("BOTTOMPADDING", (0, 0), (-1, -1), 3),
        ]))
        
        elements.append(table)
        
        doc.build(elements)
        return filepath


# Convenience function
def export_report(
    report: "FinancialStatement",
    filepath: Union[str, Path],
    format: str = "csv",
) -> Path:
    """
    Export a financial report to the specified format.
    
    Args:
        report: The financial statement to export.
        filepath: The path to save the exported file.
        format: The export format (csv, json, excel, pdf).
        
    Returns:
        The path to the exported file.
        
    Raises:
        ValueError: If the format is not supported.
    """
    exporters = {
        "csv": CSVExporter,
        "json": JSONExporter,
        "excel": ExcelExporter,
        "xlsx": ExcelExporter,
        "pdf": PDFExporter,
    }
    
    format = format.lower()
    if format not in exporters:
        raise ValueError(
            f"Unsupported format: {format}. "
            f"Supported formats: {', '.join(exporters.keys())}"
        )
    
    exporter = exporters[format](report)
    return exporter.export(filepath)
